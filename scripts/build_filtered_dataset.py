#!/usr/bin/env python3
import argparse
import csv
import hashlib
import io
import json
from io import BytesIO
import re
import sqlite3
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import requests
import trafilatura
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency in local dev only
    PdfReader = None


URL_RE = re.compile(r'https?://\S+')
MSG_START_RE = re.compile(
    r'^\u200e?\[(?P<date>\d{1,2}/\d{1,2}/\d{2}),\s(?P<time>\d{1,2}:\d{2}:\d{2}\s?[AP]M)\]\s(?P<sender>[^:]+):\s?(?P<body>.*)$'
)
ATTACH_RE = re.compile(r'<attached:\s*([^>]+)>')
TRACKING_PARAMS = {
    'utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content',
    'fbclid', 'gclid', 'mc_cid', 'mc_eid'
}
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
SOCIAL_DOMAINS = {
    'facebook.com', 'www.facebook.com', 'm.facebook.com',
    'instagram.com', 'www.instagram.com',
    'x.com', 'www.x.com', 'twitter.com', 'www.twitter.com',
    'tiktok.com', 'www.tiktok.com',
    'linkedin.com', 'www.linkedin.com',
    'youtube.com', 'www.youtube.com', 'youtu.be',
    'reddit.com', 'www.reddit.com',
}
JUNK_TEXT_MARKERS = [
    'skip to content',
    'new chat',
    'search chats',
    'privacy policy',
    'log in',
    'sign up',
    'cookie',
    'terms of use',
    'accept all',
    'subscribe',
    'enable javascript',
    'press and hold',
]
BLOCK_MARKERS = [
    'enable javascript',
    'enable cookies',
    'please verify you are human',
    'verify you are human',
    'captcha',
    'access denied',
    'temporarily blocked',
    'cloudflare',
    'press and hold',
    'checking if the site connection is secure',
]
SUMMARY_PROMPT = '''You are helping a strategic foresight team process a shared article.

Use ONLY the article title and article text.
Write a concise 2-3 sentence article summary focused on the core signal and why it may matter.
Return valid JSON only:
{"summary":"..."}

Title: {title}
Article text:
{text}
'''

PRIVATE_FIELDS = [
    'record_id', 'message_time', 'sender', 'asset_type', 'link_url', 'final_url',
    'source_domain', 'source_type', 'fetch_status', 'extraction_method', 'image_path',
    'original_attachment_name', 'person_description', 'discussion_hashtags',
    'matched_taxonomy_tags', 'signal_hashtags', 'tag_origin', 'tag_review_status',
    'scraped_header', 'sub_channel_name', 'article_text_extracted', 'article_summary',
    'record_quality_tier', 'summary_source', 'article_text', 'source_zip', 'message_id'
]

PUBLISHED_FIELDS = [
    'signal_id', 'message_time', 'asset_type', 'link_url', 'final_url', 'source_domain',
    'source_type', 'fetch_status', 'extraction_method', 'image_path', 'scraped_header',
    'sub_channel_name', 'article_text_extracted', 'article_summary', 'record_quality_tier',
    'summary_source', 'discussion_hashtags', 'signal_hashtags', 'tag_origin', 'tag_review_status'
]


@dataclass
class Message:
    message_id: str
    chat_name: str
    message_dt: Optional[str]
    sender: str
    body: str
    source_zip: str


def sha1(text: str) -> str:
    return hashlib.sha1(text.encode('utf-8')).hexdigest()


def strip_www(netloc: str) -> str:
    return (netloc or '').lower().removeprefix('www.')


def canonicalize_url(url: str) -> str:
    url = url.strip().rstrip(').,]')
    parsed = urlparse(url)
    query = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True)
             if k.lower() not in TRACKING_PARAMS]
    path = re.sub(r'/+', '/', parsed.path or '/')
    return urlunparse(((parsed.scheme or 'https').lower(), parsed.netloc.lower(), path.rstrip('/') or '/', '', urlencode(query), ''))


def detect_chat_name(zip_name: str) -> str:
    stem = Path(zip_name).stem
    if stem.lower().startswith('whatsapp chat - '):
        return stem[len('WhatsApp Chat - '):]
    return stem


def parse_datetime(date_str: str, time_str: str) -> Optional[str]:
    try:
        dt = datetime.strptime(f'{date_str} {time_str.replace(chr(0x202f), " ")}', '%d/%m/%y %I:%M:%S %p')
        return dt.isoformat(sep=' ')
    except Exception:
        return None


def parse_chat_text(chat_name: str, text: str, source_zip: str) -> List[Message]:
    lines = text.splitlines()
    messages = []
    current = None
    for line in lines:
        m = MSG_START_RE.match(line)
        if m:
            if current:
                messages.append(current)
            current = [m.group('date'), m.group('time'), m.group('sender').strip(), m.group('body')]
        else:
            if current:
                current[3] += '\n' + line
    if current:
        messages.append(current)

    out: List[Message] = []
    for idx, (date_str, time_str, sender, body) in enumerate(messages, start=1):
        raw = body.strip('\n')
        msg_id = sha1(f'{source_zip}|{idx}|{date_str}|{time_str}|{sender}|{raw}')
        out.append(Message(msg_id, chat_name, parse_datetime(date_str, time_str), sender, raw, source_zip))
    return out


def body_without_artifacts(body: str, urls: Sequence[str], attachments: Sequence[str]) -> str:
    cleaned = body
    for u in urls:
        cleaned = cleaned.replace(u, ' ')
    for a in attachments:
        cleaned = cleaned.replace(f'<attached: {a}>', ' ')
    cleaned = cleaned.replace('\u200e<This message was edited>', ' ')
    cleaned = re.sub(r'\u200e', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned or 'NA'


def is_image(filename: str) -> bool:
    return Path(filename).suffix.lower() in IMAGE_EXTS


def safe_slug(text: str) -> str:
    text = re.sub(r'[^A-Za-z0-9._-]+', '_', text)
    return text[:180].strip('._') or 'file'


def extract_zip_members(zip_path: Path, member_names: Sequence[str], out_dir: Path) -> Dict[str, Path]:
    extracted: Dict[str, Path] = {}
    if not member_names:
        return extracted

    with zipfile.ZipFile(zip_path) as zf:
        for member in member_names:
            if member not in zf.namelist():
                continue
            raw = zf.read(member)
            suffix = Path(member).suffix.lower() or '.bin'
            content_sha = hashlib.sha1(raw).hexdigest()[:16]
            target = out_dir / safe_slug(zip_path.stem) / f'{content_sha}{suffix}'
            target.parent.mkdir(parents=True, exist_ok=True)
            if not target.exists():
                with open(target, 'wb') as dst:
                    dst.write(raw)
            extracted[member] = target
    return extracted


def looks_blocked(text: str) -> bool:
    hay = (text or '').lower()
    return any(marker in hay for marker in BLOCK_MARKERS)


def extract_title_from_html(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, 'html.parser')
    if soup.title and soup.title.get_text(strip=True):
        return soup.title.get_text(strip=True)
    og = soup.find('meta', attrs={'property': 'og:title'})
    if og and og.get('content'):
        return og['content'].strip()
    tw = soup.find('meta', attrs={'name': 'twitter:title'})
    if tw and tw.get('content'):
        return tw['content'].strip()
    return None


def extract_meta_description_from_html(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, 'html.parser')
    candidates = [
        ('meta', {'property': 'og:description'}),
        ('meta', {'name': 'description'}),
        ('meta', {'name': 'twitter:description'}),
    ]
    for tag_name, attrs in candidates:
        tag = soup.find(tag_name, attrs=attrs)
        if tag and tag.get('content'):
            content = re.sub(r'\s+', ' ', tag['content']).strip()
            if content:
                return content
    return None


def clean_social_title(title: Optional[str], domain_label: str) -> Optional[str]:
    if not title:
        return None
    cleaned = re.sub(r'\s+', ' ', title).strip()
    cleaned = re.sub(r'\s*\|\s*LinkedIn.*$', '', cleaned, flags=re.I)
    cleaned = re.sub(r'\s*\|\s*X$', '', cleaned, flags=re.I)
    cleaned = re.sub(r'\s*on X:.*$', '', cleaned, flags=re.I)
    cleaned = re.sub(r'^\(\d+\)\s*', '', cleaned)
    cleaned = cleaned.strip(' -|')
    if not cleaned or len(cleaned) < 5:
        return None
    return cleaned


def summarize_social_metadata(title: Optional[str], description: Optional[str], domain: str, fallback_label: str) -> str:
    parts = []
    if title:
        parts.append(title)
    if description:
        desc = re.sub(r'\s+', ' ', description).strip()
        if desc and desc.lower() != (title or '').lower():
            parts.append(desc[:220].rsplit(' ', 1)[0] if len(desc) > 220 else desc)
    if parts:
        return ' — '.join(parts)
    return fallback_label or f'[{domain} post shared]'


def visible_text_from_html(html: str) -> str:
    soup = BeautifulSoup(html, 'html.parser')
    for tag in soup(['script', 'style', 'noscript', 'header', 'footer']):
        tag.extract()
    text = soup.get_text(separator=' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def fallback_title_from_url(url: str, source_type: str = 'html') -> str:
    parsed = urlparse(url)
    domain = strip_www(parsed.netloc)
    path = parsed.path.strip('/')

    typed_labels = {
        'youtube': '[YouTube video shared]',
        'x': '[X post shared]',
        'instagram': '[Instagram post shared]',
        'linkedin': '[LinkedIn post shared]',
        'spotify': '[Spotify link shared]',
        'chatgpt_share': '[ChatGPT shared page]',
        'facebook': '[Facebook link shared]',
        'tiktok': '[TikTok link shared]',
        'reddit': '[Reddit thread shared]',
    }
    if source_type in typed_labels:
        return typed_labels[source_type]

    if source_type == 'pdf':
        tail = path.split('/')[-1] if path else 'PDF document'
        tail = re.sub(r'[-_]+', ' ', tail).strip()
        return f'[PDF] {tail[:120]}' if tail else '[PDF document]'

    tail = path.split('/')[-1] if path else domain
    tail = re.sub(r'[-_]+', ' ', tail).strip()

    if not tail:
        return f'[{domain} link shared]' if domain else url

    if len(tail) < 6 or re.fullmatch(r'[A-Za-z0-9]{1,14}', tail):
        return f'[{domain} link shared]' if domain else url

    if tail.lower() in {'index', 'home', 'default', 'article', 'story', 'watch', 'post', 'status'}:
        return f'[{domain} link shared]' if domain else url

    return tail[:160]


def classify_source_type(url: str) -> str:
    parsed = urlparse(url)
    domain = strip_www(parsed.netloc)
    path = (parsed.path or '').lower()

    if path.endswith('.pdf'):
        return 'pdf'
    if domain in {'youtube.com', 'youtu.be'}:
        return 'youtube'
    if domain in {'x.com', 'twitter.com'}:
        return 'x'
    if domain in {'instagram.com'}:
        return 'instagram'
    if domain in {'linkedin.com'}:
        return 'linkedin'
    if domain in {'open.spotify.com'}:
        return 'spotify'
    if domain in {'chatgpt.com'} and '/share/' in path:
        return 'chatgpt_share'
    if domain in {'facebook.com', 'm.facebook.com'}:
        return 'facebook'
    if domain in {'tiktok.com'}:
        return 'tiktok'
    if domain in {'reddit.com'}:
        return 'reddit'
    return 'html'


def looks_like_junk_text(text: str) -> bool:
    low = (text or '').lower()
    marker_hits = sum(marker in low for marker in JUNK_TEXT_MARKERS)
    if marker_hits >= 2:
        return True
    if low.count('http') >= 3:
        return True
    words = low.split()
    if len(words) < 60:
        return True
    unique_ratio = len(set(words)) / max(len(words), 1)
    if len(words) >= 80 and unique_ratio < 0.18:
        return True
    return False


def clean_extracted_text(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    cleaned = re.sub(r'\s+', ' ', text).strip()
    if not cleaned:
        return None
    if looks_blocked(cleaned):
        return None
    if looks_like_junk_text(cleaned):
        return None
    return cleaned


def extract_pdf_text(pdf_bytes: bytes, max_pages: int = 15) -> Optional[str]:
    if PdfReader is None:
        return None
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        pages = []
        for page in reader.pages[:max_pages]:
            txt = page.extract_text() or ''
            if txt:
                pages.append(txt)
        return clean_extracted_text('\n'.join(pages))
    except Exception:
        return None


def fetch_url_data(url: str, timeout: int = 10, session: Optional[requests.Session] = None) -> dict:
    session = session or requests.Session()
    source_type = classify_source_type(url)
    out = {
        'title': None,
        'text': None,
        'final_url': url,
        'status_code': None,
        'fetch_status': 'not_attempted',
        'extraction_method': 'none',
        'source_type': source_type,
        'record_quality_tier': 'unresolved',
    }

    if source_type == 'youtube':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/124.0 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.8',
        }
        try:
            resp = session.get(url, timeout=(4, timeout), headers=headers, allow_redirects=True)
            out['status_code'] = resp.status_code
            out['final_url'] = resp.url or url

            if resp.status_code >= 400:
                out['fetch_status'] = f'http_{resp.status_code}'
                out['title'] = fallback_title_from_url(out['final_url'], source_type='youtube')
                out['extraction_method'] = 'youtube_http_error'
                out['record_quality_tier'] = 'metadata_only'
                return out

            html = resp.text or ''
            title = extract_title_from_html(html)

            if not title:
                m = re.search(r'<meta\s+name=["\']title["\']\s+content=["\']([^"\']+)["\']', html, re.I)
                if m:
                    title = m.group(1).strip()

            if not title:
                m = re.search(r'"title":"([^"\\]*(?:\\.[^"\\]*)*)"', html)
                if m:
                    title = bytes(m.group(1), 'utf-8').decode('unicode_escape').replace('\u0026', '&').strip()

            out['title'] = title or fallback_title_from_url(out['final_url'], source_type='youtube')
            out['fetch_status'] = 'metadata_only_youtube'
            out['extraction_method'] = 'youtube_html_title' if title else 'youtube_fallback'
            out['record_quality_tier'] = 'metadata_only'
            return out

        except requests.Timeout:
            out['fetch_status'] = 'timeout'
            out['title'] = fallback_title_from_url(url, source_type='youtube')
            out['extraction_method'] = 'youtube_timeout'
            out['record_quality_tier'] = 'metadata_only'
            return out
        except requests.RequestException:
            out['fetch_status'] = 'request_error'
            out['title'] = fallback_title_from_url(url, source_type='youtube')
            out['extraction_method'] = 'youtube_request_error'
            out['record_quality_tier'] = 'metadata_only'
            return out
        except Exception:
            out['fetch_status'] = 'unexpected_error'
            out['title'] = fallback_title_from_url(url, source_type='youtube')
            out['extraction_method'] = 'youtube_unexpected_error'
            out['record_quality_tier'] = 'metadata_only'
            return out

    if source_type == 'x':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/124.0 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.8',
        }
        try:
            oembed_resp = session.get(
                'https://publish.twitter.com/oembed',
                params={'url': url, 'omit_script': '1', 'dnt': 'true'},
                timeout=(4, timeout),
                headers=headers,
            )
            if oembed_resp.ok:
                payload = oembed_resp.json()
                author = re.sub(r'\s+', ' ', str(payload.get('author_name') or '')).strip()
                html_snippet = str(payload.get('html') or '')
                snippet = BeautifulSoup(html_snippet, 'html.parser').get_text(' ', strip=True)
                snippet = re.sub(r'\s+', ' ', snippet).strip()
                snippet = re.sub(r'\bpic\.twitter\.com/\S+', '', snippet).strip()
                title_parts = [f'X post by {author}' if author else 'X post shared']
                if snippet:
                    title_parts.append(snippet[:220].rsplit(' ', 1)[0] if len(snippet) > 220 else snippet)
                out['title'] = ' — '.join([p for p in title_parts if p]).strip(' —')
                out['fetch_status'] = 'metadata_only_x'
                out['extraction_method'] = 'x_oembed'
                out['record_quality_tier'] = 'metadata_only'
                return out
        except Exception:
            pass

        try:
            resp = session.get(url, timeout=(4, timeout), headers=headers, allow_redirects=True)
            out['status_code'] = resp.status_code
            out['final_url'] = resp.url or url
            if resp.status_code >= 400:
                out['fetch_status'] = f'http_{resp.status_code}'
                out['title'] = fallback_title_from_url(out['final_url'], source_type='x')
                out['extraction_method'] = 'x_http_error'
                out['record_quality_tier'] = 'metadata_only'
                return out

            html = resp.text or ''
            title = clean_social_title(extract_title_from_html(html), 'x')
            description = extract_meta_description_from_html(html)
            if title or description:
                out['title'] = summarize_social_metadata(title, description, 'x.com', '[X post shared]')
                out['fetch_status'] = 'metadata_only_x'
                out['extraction_method'] = 'x_html_metadata'
                out['record_quality_tier'] = 'metadata_only'
                return out
        except Exception:
            pass

        out['fetch_status'] = 'metadata_only_x'
        out['title'] = fallback_title_from_url(url, source_type='x')
        out['extraction_method'] = 'x_fallback'
        out['record_quality_tier'] = 'metadata_only'
        return out

    if source_type == 'linkedin':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/124.0 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.8',
        }
        try:
            resp = session.get(url, timeout=(4, timeout), headers=headers, allow_redirects=True)
            out['status_code'] = resp.status_code
            out['final_url'] = resp.url or url
            if resp.status_code >= 400:
                out['fetch_status'] = f'http_{resp.status_code}'
                out['title'] = fallback_title_from_url(out['final_url'], source_type='linkedin')
                out['extraction_method'] = 'linkedin_http_error'
                out['record_quality_tier'] = 'metadata_only'
                return out

            html = resp.text or ''
            title = clean_social_title(extract_title_from_html(html), 'linkedin')
            description = extract_meta_description_from_html(html)
            if title or description:
                out['title'] = summarize_social_metadata(title, description, 'linkedin.com', '[LinkedIn post shared]')
                out['fetch_status'] = 'metadata_only_linkedin'
                out['extraction_method'] = 'linkedin_html_metadata'
                out['record_quality_tier'] = 'metadata_only'
                return out
        except Exception:
            pass

        out['fetch_status'] = 'metadata_only_linkedin'
        out['title'] = fallback_title_from_url(url, source_type='linkedin')
        out['extraction_method'] = 'linkedin_fallback'
        out['record_quality_tier'] = 'metadata_only'
        return out

    if source_type in {'instagram', 'spotify', 'chatgpt_share', 'facebook', 'tiktok', 'reddit'}:
        out['fetch_status'] = f'metadata_only_{source_type}'
        out['title'] = fallback_title_from_url(url, source_type=source_type)
        out['extraction_method'] = 'metadata_only'
        out['record_quality_tier'] = 'metadata_only'
        return out

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/124.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.8',
    }

    try:
        resp = session.get(url, timeout=(4, timeout), headers=headers, allow_redirects=True)
        out['status_code'] = resp.status_code
        out['final_url'] = resp.url or url
        content_type = (resp.headers.get('content-type') or '').lower()
        out['source_type'] = classify_source_type(out['final_url'])

        if resp.status_code >= 400:
            out['fetch_status'] = f'http_{resp.status_code}'
            out['title'] = fallback_title_from_url(out['final_url'], source_type=out['source_type'])
            out['record_quality_tier'] = 'unresolved'
            return out

        if 'application/pdf' in content_type or out['final_url'].lower().endswith('.pdf'):
            out['source_type'] = 'pdf'
            out['title'] = fallback_title_from_url(out['final_url'], source_type='pdf')
            out['text'] = extract_pdf_text(resp.content)
            out['fetch_status'] = 'ok_pdf' if out['text'] else 'pdf_no_text'
            out['extraction_method'] = 'pypdf' if out['text'] else 'pdf_failed'
            out['record_quality_tier'] = 'full_article' if out['text'] else 'metadata_only'
            return out

        html = resp.text
        out['title'] = extract_title_from_html(html) or fallback_title_from_url(out['final_url'], source_type=out['source_type'])

        if looks_blocked(html) or looks_blocked(out['title'] or ''):
            out['fetch_status'] = 'blocked_anti_bot'
            out['extraction_method'] = 'metadata_only_blocked'
            out['record_quality_tier'] = 'metadata_only'
            return out

        extracted = trafilatura.extract(html, include_comments=False, include_tables=False)
        extracted = clean_extracted_text(extracted)
        if extracted:
            out['text'] = extracted
            out['fetch_status'] = 'ok_html'
            out['extraction_method'] = 'trafilatura'
            out['record_quality_tier'] = 'full_article'
            return out

        visible = clean_extracted_text(visible_text_from_html(html))
        if visible:
            out['text'] = visible
            out['fetch_status'] = 'ok_html_fallback'
            out['extraction_method'] = 'beautifulsoup_visible_text'
            out['record_quality_tier'] = 'full_article'
            return out

        out['fetch_status'] = 'html_no_text'
        out['extraction_method'] = 'html_failed'
        out['record_quality_tier'] = 'metadata_only'
        return out
    except requests.Timeout:
        out['fetch_status'] = 'timeout'
    except requests.RequestException:
        out['fetch_status'] = 'request_error'
    except Exception:
        out['fetch_status'] = 'unexpected_error'

    out['title'] = fallback_title_from_url(out['final_url'], source_type=out['source_type'])
    out['record_quality_tier'] = 'unresolved'
    return out


def heuristic_summary_from_article(title: str, article_text: str) -> str:
    title = (title or '').strip()
    article_text = re.sub(r'\s+', ' ', (article_text or '')).strip()
    if not article_text:
        return 'NA'
    snippet = article_text[:420].rsplit(' ', 1)[0]
    if title:
        return f'{title}: {snippet}…'
    return snippet + ('…' if len(article_text) > len(snippet) else '')


def fallback_signal_summary(title: str, discussion: str, domain: str, fetch_status: str) -> str:
    discussion = normalize_space(discussion or '')
    title = normalize_space(title or '')
    domain = domain or 'unknown source'
    fetch_status = fetch_status or 'unknown_status'

    if discussion and discussion != 'NA':
        snippet = discussion[:260].rsplit(' ', 1)[0]
        return f'Shared from {domain}. Discussion context: {snippet}.'
    if title and title != 'NA':
        return f'Shared from {domain}. Available metadata: {title}.'
    return f'Shared from {domain}. No article text was extracted ({fetch_status}).'


def call_ollama_summary(model: str, title: str, text: str, timeout: int = 120) -> Optional[str]:
    payload = {
        'model': model,
        'prompt': SUMMARY_PROMPT.format(title=title or 'NA', text=(text or 'NA')[:12000]),
        'format': 'json',
        'stream': False,
        'options': {'temperature': 0.2},
    }
    try:
        resp = requests.post('http://localhost:11434/api/generate', json=payload, timeout=timeout)
        resp.raise_for_status()
        raw = resp.json().get('response', '').strip()
        parsed = json.loads(raw)
        summary = (parsed.get('summary') or '').strip() if isinstance(parsed, dict) else ''
        return summary or None
    except Exception:
        return None


def normalize_space(text: str) -> str:
    return re.sub(r'\s+', ' ', text).strip()


def normalize_text_for_match(text: str) -> str:
    text = (text or '').lower()
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return normalize_space(text)


def ensure_tag(text: str) -> str:
    text = str(text or '').strip()
    if not text:
        return ''
    return text if text.startswith('#') else f'#{text}'


def tag_key(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(text or '').lower())


def extract_inline_hashtags(text: str) -> List[str]:
    if not text or text == 'NA':
        return []
    return [t for t in re.findall(r'#([A-Za-z0-9_\-/]+)', text)]


def load_tag_taxonomy(path: Optional[Path]) -> Tuple[List[dict], Dict[str, str]]:
    if not path or not path.exists():
        return [], {}

    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, dict) and 'tags' in data:
        raw_entries = data['tags']
    elif isinstance(data, dict):
        raw_entries = [{'canonical': k, **(v if isinstance(v, dict) else {})} for k, v in data.items()]
    elif isinstance(data, list):
        raw_entries = data
    else:
        raw_entries = []

    entries: List[dict] = []
    alias_map: Dict[str, str] = {}
    for item in raw_entries:
        canonical = ensure_tag(item.get('canonical') or item.get('tag') or item.get('name'))
        if not canonical:
            continue
        aliases = item.get('aliases', []) or []
        keywords = item.get('keywords', []) or []
        domains = [strip_www(d) for d in item.get('domains', []) or []]

        entry = {
            'canonical': canonical,
            'canonical_key': tag_key(canonical),
            'aliases': [str(a).strip() for a in aliases if str(a).strip()],
            'keywords': [str(k).strip() for k in keywords if str(k).strip()],
            'domains': domains,
        }
        entries.append(entry)

        for alias in [canonical, *aliases]:
            alias_map[tag_key(alias)] = canonical

    return entries, alias_map


def phrase_in_text(phrase: str, normalized_text: str) -> bool:
    needle = normalize_text_for_match(phrase)
    if not needle or not normalized_text:
        return False
    if len(needle) <= 2:
        return bool(re.search(rf'(^| ){re.escape(needle)}( |$)', normalized_text))
    return f' {needle} ' in f' {normalized_text} '


def normalize_discussion_tags(explicit_tags: Sequence[str], alias_map: Dict[str, str]) -> List[str]:
    out: List[str] = []
    for tag in explicit_tags:
        raw = ensure_tag(tag)
        canonical = alias_map.get(tag_key(raw)) or raw
        if canonical not in out:
            out.append(canonical)
    return out


def match_taxonomy_tags(texts: Sequence[str], taxonomy: Sequence[dict], source_domain: str = '') -> List[str]:
    combined = normalize_text_for_match(' '.join([t for t in texts if t and t != 'NA']))
    matches: List[str] = []
    if not combined or not taxonomy:
        return matches

    source_domain = strip_www(source_domain)
    for entry in taxonomy:
        if entry['domains'] and source_domain and source_domain not in entry['domains']:
            continue
        phrases = [entry['canonical'], *entry['aliases'], *entry['keywords']]
        if any(phrase_in_text(phrase, combined) for phrase in phrases):
            matches.append(entry['canonical'])
    return matches


def determine_signal_tags(discussion_text: str, title: str, article_text: str, source_domain: str,
                          taxonomy: Sequence[dict], alias_map: Dict[str, str]) -> Tuple[List[str], List[str], str, str, List[str]]:
    explicit = normalize_discussion_tags(extract_inline_hashtags(discussion_text), alias_map)
    taxonomy_matches = match_taxonomy_tags([discussion_text, title, article_text, source_domain], taxonomy, source_domain=source_domain)

    if explicit:
        final_tags = explicit
        tag_origin = 'discussion_explicit'
    elif taxonomy_matches:
        final_tags = taxonomy_matches
        if article_text and article_text != 'NA':
            tag_origin = 'taxonomy_match_article'
        elif discussion_text and discussion_text != 'NA':
            tag_origin = 'taxonomy_match_discussion'
        else:
            tag_origin = 'taxonomy_match_metadata'
    else:
        final_tags = []
        tag_origin = 'none'

    review_status = 'ok' if final_tags else 'needs_review'
    return explicit, taxonomy_matches, tag_origin, review_status, final_tags


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    DROP TABLE IF EXISTS processed_signals_private;
    CREATE TABLE processed_signals_private (
        record_id TEXT PRIMARY KEY,
        message_time TEXT,
        sender TEXT,
        asset_type TEXT,
        link_url TEXT,
        final_url TEXT,
        source_domain TEXT,
        source_type TEXT,
        fetch_status TEXT,
        extraction_method TEXT,
        image_path TEXT,
        original_attachment_name TEXT,
        person_description TEXT,
        discussion_hashtags TEXT,
        matched_taxonomy_tags TEXT,
        signal_hashtags TEXT,
        tag_origin TEXT,
        tag_review_status TEXT,
        scraped_header TEXT,
        sub_channel_name TEXT,
        article_text_extracted TEXT,
        article_summary TEXT,
        record_quality_tier TEXT,
        summary_source TEXT,
        article_text TEXT,
        source_zip TEXT,
        message_id TEXT
    );
    """)
    conn.commit()


def collect_records(input_dir: Path, image_dir: Path) -> List[dict]:
    all_records: List[dict] = []
    for zip_path in sorted(input_dir.glob('*.zip')):
        chat_name = detect_chat_name(zip_path.name)
        with zipfile.ZipFile(zip_path) as zf:
            names = set(zf.namelist())
            txt_name = '_chat.txt' if '_chat.txt' in names else next((n for n in names if n.endswith('_chat.txt')), None)
            if not txt_name:
                continue
            text = zf.read(txt_name).decode('utf-8-sig', errors='replace')
            messages = parse_chat_text(chat_name, text, zip_path.name)

        needed_images: List[str] = []
        parsed = []
        for msg in messages:
            original_urls = URL_RE.findall(msg.body)
            urls = [canonicalize_url(u) for u in original_urls]
            attachments = ATTACH_RE.findall(msg.body)
            image_attachments = [a for a in attachments if is_image(a)]
            if not urls and not image_attachments:
                continue
            description = body_without_artifacts(msg.body, original_urls, attachments)
            parsed.append((msg, urls, image_attachments, description))
            needed_images.extend(image_attachments)

        image_map = extract_zip_members(zip_path, needed_images, image_dir)

        for msg, urls, image_attachments, description in parsed:
            for idx, url in enumerate(urls, start=1):
                all_records.append({
                    'record_id': sha1(f'{msg.message_id}|url|{idx}|{url}'),
                    'message_time': msg.message_dt or 'NA',
                    'sender': msg.sender,
                    'asset_type': 'link',
                    'link_url': url,
                    'final_url': url,
                    'source_domain': strip_www(urlparse(url).netloc),
                    'source_type': classify_source_type(url),
                    'fetch_status': 'not_attempted',
                    'extraction_method': 'none',
                    'image_path': None,
                    'original_attachment_name': None,
                    'person_description': description,
                    'discussion_hashtags': 'NA',
                    'matched_taxonomy_tags': 'NA',
                    'signal_hashtags': 'NA',
                    'tag_origin': 'none',
                    'tag_review_status': 'needs_review',
                    'scraped_header': 'NA',
                    'sub_channel_name': msg.chat_name,
                    'article_text_extracted': 'no',
                    'article_summary': 'NA',
                    'record_quality_tier': 'unresolved',
                    'summary_source': 'none',
                    'article_text': None,
                    'source_zip': msg.source_zip,
                    'message_id': msg.message_id,
                })
            for idx, att in enumerate(image_attachments, start=1):
                img_path = image_map.get(att)
                all_records.append({
                    'record_id': sha1(f'{msg.message_id}|image|{idx}|{att}'),
                    'message_time': msg.message_dt or 'NA',
                    'sender': msg.sender,
                    'asset_type': 'image',
                    'link_url': None,
                    'final_url': None,
                    'source_domain': 'NA',
                    'source_type': 'image',
                    'fetch_status': 'not_applicable',
                    'extraction_method': 'not_applicable',
                    'image_path': str(img_path) if img_path else 'NA',
                    'original_attachment_name': Path(att).name,
                    'person_description': description,
                    'discussion_hashtags': 'NA',
                    'matched_taxonomy_tags': 'NA',
                    'signal_hashtags': 'NA',
                    'tag_origin': 'none',
                    'tag_review_status': 'needs_review',
                    'scraped_header': 'NA',
                    'sub_channel_name': msg.chat_name,
                    'article_text_extracted': 'NA',
                    'article_summary': 'NA',
                    'record_quality_tier': 'metadata_only',
                    'summary_source': 'none',
                    'article_text': None,
                    'source_zip': msg.source_zip,
                    'message_id': msg.message_id,
                })
    return all_records


def enrich_records(records: List[dict], taxonomy: Sequence[dict], alias_map: Dict[str, str],
                   use_ollama: bool = False, ollama_model: str = 'llama3.1:8b',
                   skip_fetch: bool = False) -> List[dict]:
    session = requests.Session()
    for rec in records:
        if rec['asset_type'] == 'link':
            meta = {
                'title': None,
                'text': None,
                'final_url': rec['link_url'],
                'fetch_status': 'skipped_by_flag' if skip_fetch else 'not_attempted',
                'extraction_method': 'none',
                'source_type': rec['source_type'],
            }
            if not skip_fetch:
                meta = fetch_url_data(rec['link_url'], session=session)

            final_url = meta.get('final_url') or rec['link_url']
            source_domain = strip_www(urlparse(final_url).netloc) or rec['source_domain']
            rec['final_url'] = final_url
            rec['source_domain'] = source_domain or 'NA'
            rec['source_type'] = meta.get('source_type') or rec['source_type']
            rec['fetch_status'] = meta.get('fetch_status') or rec['fetch_status']
            rec['extraction_method'] = meta.get('extraction_method') or rec['extraction_method']
            rec['record_quality_tier'] = meta.get('record_quality_tier') or rec.get('record_quality_tier') or 'unresolved'
            rec['scraped_header'] = meta.get('title') or fallback_title_from_url(final_url, source_type=rec['source_type'])

            article_text = (meta.get('text') or '').strip()
            rec['article_text'] = article_text or None
            rec['article_text_extracted'] = 'yes' if article_text else 'no'
            if article_text:
                summary = call_ollama_summary(ollama_model, rec['scraped_header'], article_text) if use_ollama else None
                rec['article_summary'] = summary or heuristic_summary_from_article(rec['scraped_header'], article_text)
                rec['summary_source'] = 'article_text'
            else:
                rec['article_summary'] = fallback_signal_summary(
                    rec['scraped_header'],
                    rec['person_description'],
                    rec['source_domain'],
                    rec['fetch_status'],
                )
                rec['summary_source'] = 'discussion_fallback' if rec.get('person_description') not in [None, '', 'NA'] else 'metadata_fallback'

        explicit, taxonomy_matches, tag_origin, review_status, final_tags = determine_signal_tags(
            discussion_text=rec['person_description'],
            title=rec['scraped_header'],
            article_text=rec.get('article_text') or '',
            source_domain=rec.get('source_domain') or '',
            taxonomy=taxonomy,
            alias_map=alias_map,
        )
        rec['discussion_hashtags'] = ' '.join(explicit) if explicit else 'NA'
        rec['matched_taxonomy_tags'] = ' '.join(taxonomy_matches) if taxonomy_matches else 'NA'
        rec['signal_hashtags'] = ' '.join(final_tags) if final_tags else 'NA'
        rec['tag_origin'] = tag_origin
        rec['tag_review_status'] = review_status
    return records


def write_db(conn: sqlite3.Connection, records: List[dict]) -> None:
    placeholders = ','.join(['?'] * len(PRIVATE_FIELDS))
    sql = f"INSERT INTO processed_signals_private({','.join(PRIVATE_FIELDS)}) VALUES ({placeholders})"
    for r in records:
        conn.execute(sql, tuple(r.get(field) for field in PRIVATE_FIELDS))
    conn.commit()


def write_csv(records: List[dict], out_csv: Path, fieldnames: Sequence[str]) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow({k: r.get(k) if r.get(k) not in [None, ''] else 'NA' for k in fieldnames})


def relative_or_absolute(path_text: str) -> Optional[Path]:
    if not path_text or path_text == 'NA':
        return None
    path = Path(path_text)
    if path.exists():
        return path
    candidate = Path.cwd() / path
    if candidate.exists():
        return candidate
    return None


def _xlsx_safe_image_source(image_path):
    """Return an XLImage-safe source and its original size.

    openpyxl only safely embeds gif/jpeg/png. For formats like WebP, we
    convert to in-memory PNG first so the workbook manifest and the actual
    image bytes stay aligned.
    """
    with PILImage.open(image_path) as pil:
        w, h = pil.size
        fmt = (pil.format or '').lower()
        if fmt in {'gif', 'jpeg', 'png'}:
            return str(image_path), w, h

        converted = BytesIO()
        save_img = pil
        if pil.mode not in {'RGB', 'RGBA', 'L', 'LA', 'P'}:
            save_img = pil.convert('RGBA')
        save_img.save(converted, format='PNG')
        converted.seek(0)
        return converted, w, h


def add_thumbnail(ws, cell_ref: str, image_path: str, max_px: int = 140) -> None:
    resolved = relative_or_absolute(image_path)
    if not resolved:
        ws[cell_ref] = image_path or 'NA'
        return
    try:
        source, w, h = _xlsx_safe_image_source(resolved)
        img = XLImage(source)
        scale = min(max_px / max(w, 1), max_px / max(h, 1), 1.0)
        img.width = int(w * scale)
        img.height = int(h * scale)
        ws.add_image(img, cell_ref)
    except Exception:
        ws[cell_ref] = str(resolved)


def write_private_workbook(records: List[dict], out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'signals_private'
    headers = [
        'time of message',
        'person who sent the message',
        'asset',
        'person_description',
        'discussion_hashtags',
        'signal_hashtags',
        'scraped_header',
        'sub_channel_name',
        'source_domain',
        'fetch_status',
        'article_text_extracted?',
        'article_summary',
    ]
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='D9EAD3')
    header_font = Font(bold=True)
    widths = [21, 22, 35, 40, 22, 22, 36, 26, 22, 22, 18, 60]
    for i, width in enumerate(widths, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions[get_column_letter(i)].width = width

    row_idx = 2
    for r in records:
        ws.cell(row=row_idx, column=1, value=r.get('message_time') or 'NA')
        ws.cell(row=row_idx, column=2, value=r.get('sender') or 'NA')
        if r.get('asset_type') == 'link':
            c = ws.cell(row=row_idx, column=3, value=r.get('final_url') or r.get('link_url') or 'NA')
            if r.get('final_url') or r.get('link_url'):
                c.hyperlink = r.get('final_url') or r.get('link_url')
                c.style = 'Hyperlink'
        else:
            add_thumbnail(ws, f'C{row_idx}', r.get('image_path') or 'NA')
        ws.cell(row=row_idx, column=4, value=r.get('person_description') or 'NA')
        ws.cell(row=row_idx, column=5, value=r.get('discussion_hashtags') or 'NA')
        ws.cell(row=row_idx, column=6, value=r.get('signal_hashtags') or 'NA')
        ws.cell(row=row_idx, column=7, value=r.get('scraped_header') or 'NA')
        ws.cell(row=row_idx, column=8, value=r.get('sub_channel_name') or 'NA')
        ws.cell(row=row_idx, column=9, value=r.get('source_domain') or 'NA')
        ws.cell(row=row_idx, column=10, value=r.get('fetch_status') or 'NA')
        ws.cell(row=row_idx, column=11, value=r.get('article_text_extracted') or 'NA')
        ws.cell(row=row_idx, column=12, value=r.get('article_summary') or 'NA')

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row_idx].height = 110 if r.get('asset_type') == 'image' else 72
        row_idx += 1

    ws.freeze_panes = 'A2'
    wb.save(out_xlsx)


def make_published_records(records: List[dict]) -> List[dict]:
    published = []
    for r in records:
        pub = {
            'signal_id': r['record_id'],
            'message_time': r['message_time'],
            'asset_type': r['asset_type'],
            'link_url': r['link_url'] or 'NA',
            'final_url': r['final_url'] or 'NA',
            'source_domain': r['source_domain'] or 'NA',
            'source_type': r['source_type'] or 'NA',
            'fetch_status': r['fetch_status'] or 'NA',
            'extraction_method': r['extraction_method'] or 'NA',
            'image_path': r['image_path'] or 'NA',
            'scraped_header': r['scraped_header'] or 'NA',
            'sub_channel_name': r['sub_channel_name'] or 'NA',
            'article_text_extracted': r['article_text_extracted'] or 'NA',
            'article_summary': r['article_summary'] or 'NA',
            'record_quality_tier': r.get('record_quality_tier') or 'unresolved',
            'summary_source': r.get('summary_source') or 'unknown',
            'discussion_hashtags': r['discussion_hashtags'] or 'NA',
            'signal_hashtags': r['signal_hashtags'] or 'NA',
            'tag_origin': r['tag_origin'] or 'NA',
            'tag_review_status': r['tag_review_status'] or 'NA',
        }
        published.append(pub)
    return published


def main():
    parser = argparse.ArgumentParser(description='Build private and published CSF signal datasets from WhatsApp exports.')
    parser.add_argument('--input-dir', default='data/raw_exports')
    parser.add_argument('--db-path', default='data/processed/csf_signals_private.db')
    parser.add_argument('--private-csv-path', default='data/processed/processed_signals_private.csv')
    parser.add_argument('--published-csv-path', default='data/processed/processed_signals.csv',
                        help='Sanitized CSV for the Streamlit app. This keeps the legacy filename for compatibility.')
    parser.add_argument('--private-xlsx-path', default='data/processed/processed_signals_private.xlsx')
    parser.add_argument('--image-dir', default='data/assets/images')
    parser.add_argument('--tag-taxonomy-path', default='data/config/tag_taxonomy.json')
    parser.add_argument('--skip-fetch', action='store_true', help='Skip live page fetching.')
    parser.add_argument('--use-ollama', action='store_true', help='Use a local Ollama model for article summaries only.')
    parser.add_argument('--ollama-model', default='llama3.1:8b')
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    image_dir = Path(args.image_dir)
    image_dir.mkdir(parents=True, exist_ok=True)
    Path(args.db_path).parent.mkdir(parents=True, exist_ok=True)
    Path(args.private_csv_path).parent.mkdir(parents=True, exist_ok=True)
    Path(args.published_csv_path).parent.mkdir(parents=True, exist_ok=True)
    Path(args.private_xlsx_path).parent.mkdir(parents=True, exist_ok=True)

    taxonomy, alias_map = load_tag_taxonomy(Path(args.tag_taxonomy_path))

    conn = sqlite3.connect(args.db_path)
    init_db(conn)
    records = collect_records(input_dir, image_dir)
    records = enrich_records(
        records,
        taxonomy=taxonomy,
        alias_map=alias_map,
        use_ollama=args.use_ollama,
        ollama_model=args.ollama_model,
        skip_fetch=args.skip_fetch,
    )
    published_records = make_published_records(records)

    write_db(conn, records)
    write_csv(records, Path(args.private_csv_path), PRIVATE_FIELDS)
    write_csv(published_records, Path(args.published_csv_path), PUBLISHED_FIELDS)
    write_private_workbook(records, Path(args.private_xlsx_path))
    conn.close()

    print(json.dumps({
        'records': len(records),
        'links': sum(1 for r in records if r['asset_type'] == 'link'),
        'images': sum(1 for r in records if r['asset_type'] == 'image'),
        'article_text_extracted_yes': sum(1 for r in records if r.get('article_text_extracted') == 'yes'),
        'tag_review_needed': sum(1 for r in records if r.get('tag_review_status') == 'needs_review'),
        'private_csv': args.private_csv_path,
        'published_csv': args.published_csv_path,
        'private_xlsx': args.private_xlsx_path,
        'db': args.db_path,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
